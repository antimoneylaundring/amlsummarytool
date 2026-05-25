import streamlit as st
import pandas as pd
from datetime import timedelta
from io import BytesIO
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import streamlit.components.v1 as components
import os
from dotenv import load_dotenv

load_dotenv()

st.markdown("""
<style>
.block-container {
    padding-top: 1rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    max-width: 100% !important;
}
iframe { width: 100% !important; display: block !important; margin: 0 !important; }
html, body { margin: 0; padding: 0; width: 100% !important; overflow-x: hidden !important; }
.excel-table { width: 100% !important; table-layout: fixed !important; }
</style>
""", unsafe_allow_html=True)


# ================= HELPER FUNCTIONS =================
def find_column(cols, keys):
    for c in cols:
        cc = c.lower().replace(" ", "").replace("_", "")
        for k in keys:
            if k in cc:
                return c
    return None


@st.cache_resource
def get_db_engine():
    try:
        db_url = os.getenv("DB_URL")
        if not db_url:
            st.error("DB_URL not found in environment variables")
            return None
        return create_engine(db_url)
    except Exception as e:
        st.error(f"Database connection failed: {e}")
        return None


def chunk_list(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def count_new_upis_for_date(engine, upi_array, cutoff_date):
    if not upi_array:
        return 0
    total_new = 0
    try:
        with engine.connect() as conn:
            for chunk in chunk_list(upi_array, 3000):
                result = conn.execute(
                    text("SELECT missing_count FROM count_new_upi(:p_upi_array, :p_cutoff_date)"),
                    {"p_upi_array": chunk, "p_cutoff_date": cutoff_date}
                ).fetchone()
                if result is not None:
                    total_new += result[0]
        return total_new
    except Exception as e:
        st.error(f"Error counting new UPIs: {e}")
        return 0


def count_new_banks_for_date(engine, bank_array, cutoff_date):
    if not bank_array:
        return 0
    total_new = 0
    try:
        with engine.connect() as conn:
            for chunk in chunk_list(bank_array, 3000):
                result = conn.execute(
                    text("SELECT missing_count FROM count_new_bank(:p_bank_array, :p_cutoff_date)"),
                    {"p_bank_array": chunk, "p_cutoff_date": cutoff_date}
                ).fetchone()
                if result is not None:
                    total_new += result[0]
        return total_new
    except Exception as e:
        st.error(f"Error counting new banks: {e}")
        return 0


def clean_val(x):
    if pd.isna(x):
        return None
    return str(x).strip().lower().replace(" ", "")


def clean_bank_val(x):
    if pd.isna(x):
        return None
    return str(x).strip()


def process_df(date_df, engine, cutoff_date):
    """Helper to compute UPI & Bank stats for a given sub-dataframe."""
    if date_df.empty:
        return 0, 0, 0, 0, 0, 0

    date_df = date_df.copy()
    date_df["Upi_vpa_clean"] = date_df["Upi_vpa"].astype(str).str.strip().str.lower().str.replace(" ", "")
    date_df["Bank_acc_clean"] = date_df["Bank_account_number"].apply(clean_bank_val)

    upi_sub = date_df[date_df["Upi_bank_account_wallet"].astype(str).str.strip().str.upper() == "UPI"]
    total_upi = len(upi_sub)
    unique_upi_list = upi_sub["Upi_vpa_clean"].dropna().unique().tolist()
    unique_upi = len(unique_upi_list)
    new_upi = count_new_upis_for_date(engine, unique_upi_list, cutoff_date) if unique_upi_list else 0

    bank_sub = date_df[date_df["Upi_bank_account_wallet"].astype(str).str.strip() == "Bank Account"]
    total_bank = len(bank_sub)
    unique_bank_list = bank_sub["Bank_acc_clean"].dropna().unique().tolist()
    unique_bank = len(unique_bank_list)
    new_bank = count_new_banks_for_date(engine, unique_bank_list, cutoff_date) if unique_bank_list else 0

    return total_upi, unique_upi, new_upi, total_bank, unique_bank, new_bank


# ================= EXCEL EXPORT =================
# ✅ Function signature update karo
def build_excel(summary_df, multiple_summary_df, freelancer_summary_df, 
                daily_summary_df=None, crypto_summary_df=None, notFound_summary_df=None,
                website_searching=None, credential_making=None, 
                remark_checking=None, upi_fraud=None, remark_col=None):
    wb = Workbook()

    header_fill   = PatternFill("solid", fgColor="CBD5E1")
    subheader_fill = PatternFill("solid", fgColor="E2E8F0")
    green_fill    = PatternFill("solid", fgColor="CFE8B0")
    crypto_fill   = PatternFill("solid", fgColor="EBB7AC")
    notfound_fill = PatternFill("solid", fgColor="FFFFFF")
    ab_fill       = PatternFill("solid", fgColor="E9DCC2")   # AB row — beige
    bold   = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    def set_header(ws, row, col, value, rowspan=1, colspan=1, fill=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = bold
        cell.alignment = center
        if fill:
            cell.fill = fill
        if rowspan > 1 or colspan > 1:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row + rowspan - 1, end_column=col + colspan - 1
            )

    # ============ SHEET 1: UPI & Bank Summary ============
    ws1 = wb.active
    ws1.title = "UPI & Bank Summary"

    set_header(ws1, 1, 1, "UPI, Bank & Website Report", colspan=13, fill=header_fill)
    set_header(ws1, 2, 1, "Date", rowspan=2, fill=header_fill)
    set_header(ws1, 2, 2, "Total", rowspan=2, fill=header_fill)
    set_header(ws1, 2, 3, "UPI", colspan=5, fill=header_fill)
    set_header(ws1, 2, 8, "Bank", colspan=5, fill=header_fill)
    set_header(ws1, 2, 13, "Unique Website", rowspan=2, fill=header_fill)

    for col, label in enumerate(["Total", "Unique", "%", "New", "%"], start=3):
        set_header(ws1, 3, col, label, fill=subheader_fill)
    for col, label in enumerate(["Total", "Unique", "%", "New", "%"], start=8):
        set_header(ws1, 3, col, label, fill=subheader_fill)

    for r, row in enumerate(summary_df.itertuples(), start=4):
        ws1.append([
            str(row.Date), row.Total,
            row.UPI_Total, row.UPI_Unique, row.UPI_pct, row.UPI_New, row.UPI_New_pct,
            row.Bank_Total, row.Bank_Unique, row.Bank_pct, row.Bank_New, row.Bank_New_pct,
            row.unique_website
        ])
        for col in range(1, 14):
            ws1.cell(row=r, column=col).alignment = center

    for col_idx, width in enumerate([12, 7, 7, 7, 6, 6, 7, 7, 7, 6, 6, 7, 10], start=1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # ============ SHEET 2: Multiple User Summary ============
    ws2 = wb.create_sheet("Multiple User Summary")

    dates = multiple_summary_df["Date"].unique()
    date_label = str(dates[-1]) if len(dates) > 0 else "N/A"

    set_header(ws2, 1, 1, f"Multiple User's Counts ({date_label})", colspan=8, fill=green_fill)
    set_header(ws2, 2, 1, "Name", rowspan=2, fill=green_fill)
    set_header(ws2, 2, 2, "Total", rowspan=2, fill=green_fill)
    set_header(ws2, 2, 3, "Web", rowspan=2, fill=green_fill)
    set_header(ws2, 2, 4, "App", rowspan=2, fill=green_fill)
    set_header(ws2, 2, 5, "Unique UPI", colspan=2, fill=green_fill)
    set_header(ws2, 2, 7, "New UPI", colspan=2, fill=green_fill)
    for col, label in enumerate(["Count", "%", "Count", "%"], start=5):
        set_header(ws2, 3, col, label, fill=green_fill)

    for r, row in enumerate(multiple_summary_df.itertuples(), start=4):
        ws2.append([
            row.Input_user, row.Total, row.Web_Count, row.App_Count,
            row.Unique_UPI_Count, row.Unique_UPI_pct,
            row.New_UPI_Count, row.New_UPI_pct
        ])
        for col in range(1, 9):
            ws2.cell(row=r, column=col).alignment = center

    # Totals row — skip AB rows
    total_row = r + 1
    _num_df = multiple_summary_df[multiple_summary_df["Total"] != "AB"]
    total_total  = pd.to_numeric(_num_df["Total"],           errors="coerce").sum()
    total_web    = pd.to_numeric(_num_df["Web_Count"],       errors="coerce").sum()
    total_app    = pd.to_numeric(_num_df["App_Count"],       errors="coerce").sum()
    total_unique = pd.to_numeric(_num_df["Unique_UPI_Count"],errors="coerce").sum()
    total_new    = pd.to_numeric(_num_df["New_UPI_Count"],   errors="coerce").sum()
    total_unique_pct = f"{(total_unique / total_total * 100):.0f}%" if total_total else "0%"
    total_new_pct = f"{(total_new / total_unique * 100):.0f}%" if total_unique else "0%"

    ws2.append(["Total", total_total, total_web, total_app, total_unique, total_unique_pct, total_new, total_new_pct])
    for col in range(1, 9):
        cell = ws2.cell(row=total_row, column=col)
        cell.font = bold
        cell.fill = green_fill
        cell.alignment = center

    for col_idx, width in enumerate([25, 8, 12, 8, 12, 8, 12, 8], start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # ============ SHEET 3: Employee, Intern & Freelancer ============
    ws3 = wb.create_sheet("Emp Intern Freelancer")

    set_header(ws3, 1, 1, "Employee, Intern & Freelancer Summary", colspan=9, fill=green_fill)
    set_header(ws3, 2, 1, "User", rowspan=2, fill=green_fill)
    set_header(ws3, 2, 2, "Date", rowspan=2, fill=green_fill)
    set_header(ws3, 2, 3, "UPI", colspan=3, fill=green_fill)
    set_header(ws3, 2, 6, "Bank", colspan=3, fill=green_fill)
    set_header(ws3, 2, 9, "Total", rowspan=2, fill=green_fill)  # ✅ Total column header

    for col, label in enumerate(["Total UPI", "Unique UPI", "New UPI"], start=3):
        set_header(ws3, 3, col, label, fill=green_fill)
    for col, label in enumerate(["Total Bank", "Unique Bank", "New Bank"], start=6):
        set_header(ws3, 3, col, label, fill=green_fill)

    row_idx = 4
    for date in sorted(freelancer_summary_df["Date"].unique()):
        day_df = freelancer_summary_df[freelancer_summary_df["Date"] == date]
        for user_type in ["Employee", "INT", "Freelancer"]:
            r = day_df[day_df["User_Type"] == user_type]
            if not r.empty:
                r = r.iloc[0]
                t_upi  = r["Total_UPI"]
                u_upi  = r["Unique_UPI"]
                n_upi  = r["New_UPI"]
                t_bank = r["Total_Bank"]
                u_bank = r["Unique_Bank"]
                n_bank = r["New_Bank"]
            else:
                t_upi = u_upi = n_upi = t_bank = u_bank = n_bank = 0

            # ✅ Total = Total_UPI + Total_Bank
            row_total = int(t_upi) + int(t_bank)

            ws3.append([
                user_type, str(date),
                t_upi, u_upi, n_upi,
                t_bank, u_bank, n_bank,
                row_total  # ✅ Total column
            ])
            for col in range(1, 10):
                ws3.cell(row=row_idx, column=col).alignment = center
            row_idx += 1

        # Totals per date
        totals = day_df.sum(numeric_only=True)
        t_upi_sum  = int(totals.get("Total_UPI", 0))
        t_bank_sum = int(totals.get("Total_Bank", 0))

        ws3.append([
            "Total", "NA",
            t_upi_sum,
            int(totals.get("Unique_UPI", 0)),
            int(totals.get("New_UPI", 0)),
            t_bank_sum,
            int(totals.get("Unique_Bank", 0)),
            int(totals.get("New_Bank", 0)),
            t_upi_sum + t_bank_sum  # ✅ Total row ka Total
        ])
        for col in range(1, 10):
            cell = ws3.cell(row=row_idx, column=col)
            cell.font = bold
            cell.alignment = center
        row_idx += 1

    for col_idx, width in enumerate([12, 12, 10, 10, 10, 10, 12, 10, 10], start=1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    # ============ SHEET 4: Daily Summary ============
    if daily_summary_all is not None and len(daily_summary_all) > 0:
        ws4 = wb.create_sheet("Daily Summary")

        daily_dates_xl = sorted(daily_summary_all.keys())
        daily_date_xl_str = ", ".join(str(d) for d in daily_dates_xl) if daily_dates_xl else "N/A"

        first_date = daily_dates_xl[0]
        base_cols = list(daily_summary_all[first_date].columns)

        extra_cols = []
        if website_searching:
            extra_cols = ["Website Searching", "Remarks Checking", "Credentials", "UPI Fraud", "Remark"]

        daily_cols = base_cols + extra_cols
        total_cols = len(daily_cols)

        # ✅ Row 1: Title
        set_header(ws4, 1, 1, "Daily wise user summary", colspan=total_cols, fill=header_fill)

        # ✅ Row 2: Date
        set_header(ws4, 2, 1, daily_date_xl_str, colspan=total_cols, fill=subheader_fill)

        # ✅ Row 3: Section headers — Name, Insertion, Quality Check, Website Monitoring, SM Scrawling, Remark
        # Name column — rowspan 2
        set_header(ws4, 3, 1, "Name", rowspan=2, fill=header_fill)

        # Insertion — Daily Cases to Error = 9 columns (col 2 to 10)
        set_header(ws4, 3, 2, "Insertion", colspan=9, fill=header_fill)

        # Quality Check — Total QC = 1 column (col 11)
        set_header(ws4, 3, 11, "Quality Check", colspan=1, fill=subheader_fill)

        current_col = 12
        if extra_cols:
            # Website Monitoring — Website Searching, Remarks Checking, Credentials = 3 cols
            set_header(ws4, 3, current_col, "Website Monitoring", colspan=3, fill=subheader_fill)
            # SM Scrawling — UPI Fraud = 1 col
            set_header(ws4, 3, current_col + 3, "SM Scrawling", colspan=1, fill=subheader_fill)
            # Remark = 1 col
            set_header(ws4, 3, current_col + 4, "Remark", rowspan=2, fill=subheader_fill)

        # ✅ Row 4: Column headers
        col_labels = [
            "Daily Cases", "Multiple Cases", "Not Found", "App",
            "WA/TG Case", "Crypto Cases", "International Cases", "Total Case", "Error",
            "Total QC"
        ]
        for ci, label in enumerate(col_labels, start=2):
            set_header(ws4, 4, ci, label, fill=subheader_fill)

        if extra_cols:
            extra_labels = ["Website Searching", "Remarks Checking", "Credentials", "UPI Fraud"]
            for ci, label in enumerate(extra_labels, start=current_col):
                set_header(ws4, 4, ci, label, fill=subheader_fill)

        # ✅ get_val helper
        def get_val_excel(d, name):
            if not d:
                return "NA"
            name_lower = name.strip().lower()
            if name_lower in d:
                return d[name_lower]
            clean = name.strip()
            for prefix in ["Emp ", "INT ", "emp ", "int "]:
                if clean.startswith(prefix):
                    clean = clean[len(prefix):]
                    break
            clean_lower = clean.strip().lower()
            if clean_lower in d:
                return d[clean_lower]
            for key in d:
                if clean_lower in key or key in clean_lower:
                    return d[key]
            return "NA"

        # ✅ Data rows — row 5 se start
        row_idx = 5
        for date in sorted(daily_summary_all.keys()):
            day_df = daily_summary_all[date]

            for _, r in day_df.iterrows():
                row_data = []

                # Detect AB row (any file-1 column == "AB")
                is_ab_row = str(r.get("Daily Cases", "")).strip() == "AB"

                for col in base_cols:
                    if col in day_df.columns:
                        val = r[col]
                        if not isinstance(val, str) and pd.isna(val):
                            row_data.append("NA")
                        else:
                            row_data.append(val)
                    else:
                        row_data.append("NA")

                if extra_cols:
                    name = str(r["Name"]).strip()
                    if name == "Total":
                        row_data += ["NA"] * len(extra_cols)
                    elif is_ab_row:
                        # AB rows: mark all extra cols as AB too
                        row_data += ["AB"] * len(extra_cols)
                    else:
                        ws  = get_val_excel(website_searching, name)
                        rc  = get_val_excel(remark_checking, name)
                        cr  = get_val_excel(credential_making, name)
                        uf  = get_val_excel(upi_fraud, name)
                        rem = get_val_excel(remark_col, name)

                        ws  = "NA" if ws  == 0 or ws  == "NA" else ws
                        rc  = "NA" if rc  == 0 or rc  == "NA" else rc
                        cr  = "NA" if cr  == 0 or cr  == "NA" else cr
                        uf  = "NA" if uf  == 0 or uf  == "NA" else uf

                        row_data += [ws, rc, cr, uf, rem]

                ws4.append(row_data)
                is_total = str(row_data[0]) == "Total"
                for ci in range(1, total_cols + 1):
                    cell = ws4.cell(row=row_idx, column=ci)
                    cell.alignment = center
                    if is_total:
                        cell.font  = bold
                        cell.fill  = subheader_fill
                    elif is_ab_row:
                        cell.font  = Font(bold=True)
                        cell.fill  = ab_fill
                row_idx += 1

        col_widths = [22] + [13] * (total_cols - 1)
        for ci, width in enumerate(col_widths, start=1):
            ws4.column_dimensions[get_column_letter(ci)].width = width

    # ============ SHEET 5: Crypto Summary ============
    if crypto_summary_df is not None and not crypto_summary_df.empty:
        ws5 = wb.create_sheet("Crypto Summary")

        dates = crypto_summary_df["Date"].unique()
        date_label = str(dates[-1]) if len(dates) > 0 else "N/A"

        set_header(ws5, 1, 1, f"Crypto User's Counts ({date_label})", colspan=4, fill=crypto_fill)
        set_header(ws5, 2, 1, "Name", fill=crypto_fill)
        set_header(ws5, 2, 2, "Total", fill=crypto_fill)
        set_header(ws5, 2, 3, "Crypto App", fill=crypto_fill)
        set_header(ws5, 2, 4, "Crypto Web", fill=crypto_fill)

        display_crypto = (
            crypto_summary_df.groupby("Name")[["Total", "Crypto App", "Crypto Web"]]
            .sum()
            .reset_index()
        )

        row_idx = 3
        for _, row in display_crypto.iterrows():  # ✅ itertuples ki jagah iterrows
            ws5.append([row["Name"], row["Total"], row["Crypto App"], row["Crypto Web"]])
            for col in range(1, 5):
                ws5.cell(row=row_idx, column=col).alignment = center
            row_idx += 1

        # Totals row
        total_website = display_crypto["Total"].sum()
        total_app     = display_crypto["Crypto App"].sum()
        total_web     = display_crypto["Crypto Web"].sum()

        ws5.append(["Total", total_website, total_app, total_web])
        for col in range(1, 5):
            cell = ws5.cell(row=row_idx, column=col)
            cell.font = bold
            cell.fill = crypto_fill
            cell.alignment = center

        for col_idx, width in enumerate([30, 12, 12, 12], start=1):
            ws5.column_dimensions[get_column_letter(col_idx)].width = width


    # ============ SHEET 6: Not Found Summary ============
    if notFound_summary_df is not None and not notFound_summary_df.empty:
        ws6 = wb.create_sheet("Not Found Summary")

        dates = notFound_summary_df["Date"].unique()
        date_label = str(dates[-1]) if len(dates) > 0 else "N/A"

        set_header(ws6, 1, 1, f"Not Found Summary ({date_label})", colspan=3, fill=notfound_fill)
        set_header(ws6, 2, 1, "Input_user", fill=notfound_fill)
        set_header(ws6, 2, 2, "Approved", fill=notfound_fill)
        set_header(ws6, 2, 3, "Rejected", fill=notfound_fill)

        display_notfound = (
            notFound_summary_df.groupby("Name")[["Approved", "Rejected"]]
            .sum()
            .reset_index()
        )

        row_idx = 3
        for _, row in display_notfound.iterrows():  # ✅ itertuples ki jagah iterrows
            ws6.append([row["Name"], row["Approved"], row["Rejected"]])
            for col in range(1, 4):
                ws6.cell(row=row_idx, column=col).alignment = center
            row_idx += 1

        # Totals row
        total_approved = display_notfound["Approved"].sum()
        total_rejected = display_notfound["Rejected"].sum()

        ws6.append(["Total", total_approved, total_rejected])
        for col in range(1, 4):
            cell = ws6.cell(row=row_idx, column=col)
            cell.font = bold
            cell.fill = notfound_fill
            cell.alignment = center

        for col_idx, width in enumerate([30, 12, 12], start=1):
            ws6.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ================= UI =================
st.title("Summary")

# ================= COPY AS IMAGE HELPER =================
def with_copy_and_edit_button(html_content, table_id="capture-table", height=550):
    """
    Wraps html_content with Copy as Image + Edit buttons.
    Edit makes Insertion cols (2-9) & Quality Check cols (10-12) inline-editable.
    Skips Name col, AB rows, and Total row.
    """
    # Convert hyphenated table_id to a valid JS identifier for function names
    js_id = table_id.replace('-', '_')
    wrapped = f"""
    <script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
    <div style="margin-bottom:8px; display:flex; align-items:center; gap:10px; flex-wrap:wrap;">
        <button id="copy-btn-{table_id}" onclick="{js_id}_copyImg()"
            style="background:#4f8ef7;color:#fff;border:none;border-radius:6px;
                   padding:7px 18px;font-size:13px;font-weight:600;cursor:pointer;
                   display:inline-flex;align-items:center;gap:6px;
                   box-shadow:0 2px 6px rgba(79,142,247,0.25);">
            \U0001f4f7 Copy as Image
        </button>
        <button id="edit-btn-{table_id}" onclick="{js_id}_startEdit()"
            style="background:#28a745;color:#fff;border:none;border-radius:6px;
                   padding:7px 18px;font-size:13px;font-weight:600;cursor:pointer;
                   display:inline-flex;align-items:center;gap:6px;
                   box-shadow:0 2px 6px rgba(40,167,69,0.25);">
            \u270f\ufe0f Edit
        </button>
        <button id="save-btn-{table_id}" onclick="{js_id}_saveEdit()"
            style="display:none;background:#fd7e14;color:#fff;border:none;border-radius:6px;
                   padding:7px 18px;font-size:13px;font-weight:600;cursor:pointer;
                   box-shadow:0 2px 6px rgba(253,126,20,0.25);">
            \U0001f4be Save
        </button>
        <button id="cancel-btn-{table_id}" onclick="{js_id}_cancelEdit()"
            style="display:none;background:#6c757d;color:#fff;border:none;border-radius:6px;
                   padding:7px 18px;font-size:13px;font-weight:600;cursor:pointer;">
            \u2716 Cancel
        </button>
        <span id="status-{table_id}" style="font-size:12px;color:#555;"></span>
    </div>
    <div id="{table_id}">
        {html_content}
    </div>
    <script>
    (function() {{
        var originals = {{}};
        var editing = false;
        var editableCols = [2,3,4,5,6,7,8,9,10,11,12,13,14,15,16];

        function getTable() {{
            return document.getElementById('{table_id}').querySelector('table');
        }}

        function setButtons(editMode) {{
            document.getElementById('edit-btn-{table_id}').style.display   = editMode ? 'none'        : 'inline-flex';
            document.getElementById('save-btn-{table_id}').style.display   = editMode ? 'inline-flex' : 'none';
            document.getElementById('cancel-btn-{table_id}').style.display = editMode ? 'inline-flex' : 'none';
            document.getElementById('copy-btn-{table_id}').disabled        = editMode;
        }}

        window.{js_id}_startEdit = function() {{
            if (editing) return;
            editing = true;
            originals = {{}};
            var table = getTable();
            if (!table) {{ alert('Table not found'); editing = false; return; }}
            var rows = table.querySelectorAll('tbody tr');
            rows.forEach(function(tr, ri) {{
                var cells = tr.querySelectorAll('td');
                if (!cells.length) return;
                var firstName = cells[0].textContent.trim();
                if (firstName === 'Total') return;
                // skip AB rows (yellow background)
                var rowStyle = tr.getAttribute('style') || '';
                if (rowStyle.indexOf('e9dcc2') !== -1) return;
                editableCols.forEach(function(ci) {{
                    var cell = cells[ci - 1];
                    if (!cell) return;
                    var key = ri + '-' + ci;
                    originals[key] = cell.innerHTML;
                    var cur = cell.textContent.trim();
                    var inp = document.createElement('input');
                    inp.type = 'text';
                    inp.value = (cur === 'NA') ? '' : cur;
                    inp.placeholder = 'NA';
                    inp.style.cssText = 'width:95%;border:1.5px solid #4f8ef7;border-radius:3px;'
                        + 'padding:2px 3px;font-size:12px;text-align:center;'
                        + 'background:#fff9e6;box-sizing:border-box;';
                    cell.innerHTML = '';
                    cell.appendChild(inp);
                }});
            }});
            setButtons(true);
        }};

        window.{js_id}_saveEdit = function() {{
            var table = getTable(); if (!table) return;
            var rows = table.querySelectorAll('tbody tr');
            rows.forEach(function(tr, ri) {{
                var cells = tr.querySelectorAll('td');
                if (!cells.length) return;
                var firstName = cells[0].textContent.trim();
                if (firstName === 'Total') return;
                var rowStyle = tr.getAttribute('style') || '';
                if (rowStyle.indexOf('e9dcc2') !== -1) return;
                editableCols.forEach(function(ci) {{
                    var cell = cells[ci - 1]; if (!cell) return;
                    var inp = cell.querySelector('input'); if (!inp) return;
                    cell.textContent = inp.value.trim() === '' ? 'NA' : inp.value.trim();
                }});
            }});
            editing = false;
            setButtons(false);
            var st = document.getElementById('status-{table_id}');
            st.textContent = '\u2705 Saved!';
            setTimeout(function() {{ st.textContent = ''; }}, 3000);
        }};

        window.{js_id}_cancelEdit = function() {{
            var table = getTable(); if (!table) return;
            var rows = table.querySelectorAll('tbody tr');
            rows.forEach(function(tr, ri) {{
                var cells = tr.querySelectorAll('td');
                editableCols.forEach(function(ci) {{
                    var cell = cells[ci - 1]; if (!cell) return;
                    var key = ri + '-' + ci;
                    if (originals[key] !== undefined) cell.innerHTML = originals[key];
                }});
            }});
            editing = false;
            setButtons(false);
        }};

        window.{js_id}_copyImg = function() {{
            var btn    = document.getElementById('copy-btn-{table_id}');
            var status = document.getElementById('status-{table_id}');
            var el     = document.getElementById('{table_id}');
            btn.disabled = true;
            btn.textContent = '\u23f3 Capturing...';
            status.textContent = '';
            html2canvas(el, {{ scale: 2, useCORS: true, backgroundColor: '#ffffff' }}).then(function(canvas) {{
                canvas.toBlob(function(blob) {{
                    var item = new ClipboardItem({{'image/png': blob}});
                    navigator.clipboard.write([item]).then(function() {{
                        status.textContent = '\u2705 Copied!';
                        btn.innerHTML = '\U0001f4f7 Copy as Image';
                        btn.disabled = false;
                        setTimeout(function() {{ status.textContent = ''; }}, 3000);
                    }}).catch(function() {{
                        status.textContent = '\u274c Copy failed (try HTTPS)';
                        btn.innerHTML = '\U0001f4f7 Copy as Image';
                        btn.disabled = false;
                    }});
                }});
            }}).catch(function() {{
                status.textContent = '\u274c Capture failed';
                btn.innerHTML = '\U0001f4f7 Copy as Image';
                btn.disabled = false;
            }});
        }};
    }})();
    </script>
    """
    return wrapped, height + 55


def with_copy_button(html_content, table_id="capture-table", height=450):
    """
    Wraps html_content with a 📷 Copy as Image button.
    Uses html2canvas to capture only the table area and copy to clipboard.
    """
    wrapped = f"""
    <script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
    <div style="margin-bottom:6px;">
        <button id="copy-btn-{table_id}" onclick="copyTableImage('{table_id}')"
            style="
                background:#4f8ef7; color:#fff; border:none; border-radius:6px;
                padding:7px 18px; font-size:13px; font-weight:600;
                cursor:pointer; display:inline-flex; align-items:center; gap:6px;
                box-shadow:0 2px 6px rgba(79,142,247,0.25);
            ">
            📷 Copy as Image
        </button>
        <span id="status-{table_id}" style="margin-left:10px; font-size:12px; color:#555;"></span>
    </div>
    <div id="{table_id}">
        {html_content}
    </div>
    <script>
    function copyTableImage(tableId) {{
        var btn = document.getElementById('copy-btn-' + tableId);
        var status = document.getElementById('status-' + tableId);
        var el = document.getElementById(tableId);
        btn.disabled = true;
        btn.textContent = '⏳ Capturing...';
        status.textContent = '';
        html2canvas(el, {{ scale: 2, useCORS: true, backgroundColor: '#ffffff' }}).then(function(canvas) {{
            canvas.toBlob(function(blob) {{
                var item = new ClipboardItem({{'image/png': blob}});
                navigator.clipboard.write([item]).then(function() {{
                    status.textContent = '✅ Copied!';
                    btn.textContent = '📷 Copy as Image';
                    btn.disabled = false;
                    setTimeout(function() {{ status.textContent = ''; }}, 3000);
                }}).catch(function(err) {{
                    status.textContent = '❌ Copy failed (try HTTPS)';
                    btn.textContent = '📷 Copy as Image';
                    btn.disabled = false;
                }});
            }});
        }}).catch(function(err) {{
            status.textContent = '❌ Capture failed';
            btn.textContent = '📷 Copy as Image';
            btn.disabled = false;
        }});
    }}
    </script>
    """
    return wrapped, height + 50  # extra height for the button row

col1, col2 = st.columns(2)

with col1:
    uploaded_file = st.file_uploader("Upload GUI Dump File", type=["xlsx", "xls", "csv"])

with col2:
    is_daily = st.session_state.get("summary_select") == "Daily Summary"
    uploaded_file_2 = st.file_uploader(
        "Upload EDTMS Dump File",
        type=["xlsx", "xls", "csv"],
        key="file2",
        disabled=not is_daily,
        help="Only enabled when 'Daily Summary' is selected"
    )

if uploaded_file:
    engine = get_db_engine()
    if not engine:
        st.error("Cannot proceed without database connection")
        st.stop()

    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str)
    else:
        df = pd.read_excel(uploaded_file, dtype=str)

    df.columns = df.columns.str.strip()
    st.success(f"File Loaded: {uploaded_file.name}")

    required_cols = [
        "Id", "Feature_type", "Approvd_status", "Input_user",
        "Inserted_date", "Website_url", "Upi_vpa",
        "Bank_account_number", "Search_for", "Upi_bank_account_wallet"
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(f"Missing columns: {missing}")
        st.stop()

    filtered_df = df[
        (df["Feature_type"].astype(str).str.strip() == "BS Money Laundering") &
        (df["Approvd_status"].astype(str).str.strip() == "1") &
        (df["Input_user"].astype(str).str.strip().str.lower() != "automated") &
        (df["Search_for"].astype(str).str.strip().isin(["App", "Web"])) &
        (df["Upi_bank_account_wallet"].astype(str).str.strip().isin(["UPI", "Bank Account"]))
    ].copy()

    if filtered_df.empty:
        st.warning("No records found after applying filters.")
        st.stop()

    # st.info(f"{len(filtered_df)} rows matched filters")

    filtered_df["Upi_vpa_clean"] = filtered_df["Upi_vpa"].apply(clean_val)
    filtered_df["Bank_acc_clean"] = filtered_df["Bank_account_number"].apply(clean_bank_val)
    filtered_df["Website_url"] = filtered_df["Website_url"].apply(clean_val)
    filtered_df["Inserted_date"] = pd.to_datetime(filtered_df["Inserted_date"], errors="coerce").dt.date

    upi_df = filtered_df[
        (filtered_df["Upi_bank_account_wallet"].astype(str).str.strip().str.lower() == "upi")
    ].copy()

    bank_df = filtered_df[
        (filtered_df["Upi_bank_account_wallet"].astype(str).str.strip() == "Bank Account")
    ].copy()

    grouped = filtered_df.groupby("Inserted_date").agg(
        website_total=('Id', 'count'),
        Total_UPI=("Upi_vpa_clean", "count"),
        Unique_UPI=("Upi_vpa_clean", pd.Series.nunique),
        unique_website=('Website_url', pd.Series.nunique)
    ).reset_index()

    bank_grouped = bank_df.groupby("Inserted_date").agg(
        Bank_Total=("Bank_acc_clean", "count"),
        Bank_Unique=("Bank_acc_clean", pd.Series.nunique)
    ).reset_index()

    grouped = grouped.merge(bank_grouped, on="Inserted_date", how="left")
    grouped[["Bank_Total", "Bank_Unique"]] = grouped[["Bank_Total", "Bank_Unique"]].fillna(0).astype(int)

    all_dates = (
        df["Inserted_date"]
        .pipe(pd.to_datetime, errors="coerce")
        .dt.date
        .dropna()
        .unique()
    )

    target_users = [
        "Emp Vidhi Satsangi",
        "Emp Shashank Sharma",
        "Emp Muskan Verma",
    ]

    # ================= DAILY SUMMARY PROCESSING (per date, like freelancer summary) =================
    daily_summary_all = {}   # { date -> DataFrame with total row }

    qc_user_col = find_column(df.columns, ["approvedby", "qcby", "qcuser"])
    video_col = find_column(df.columns, ["Video_url", "video"])

    daily_users = [
        "Emp Muskan Verma", "Emp Shashank Sharma",
        "Emp Vidhi Satsangi", "INT Bhavna Mathur",
        "INT Chandrakanta Vishwakarma", "INT Gunjan Baghel", "INT Neha Baghel","INT Nidhi Thakur", "INT Sharad Goswami", "INT Shalini Singh", "INT Pragati Singh"
    ]
    multi_users = {"Emp Vidhi Satsangi", "Emp Shashank Sharma", "Emp Muskan Verma"}

    input_col  = "Input_user"
    search_col = "Search_for"
    wallet_col = "Upi_bank_account_wallet"
    status_col = "Approvd_status"
    website_col = "Website_url"

    if qc_user_col and video_col:
        df["_date"] = pd.to_datetime(df["Inserted_date"], errors="coerce").dt.date

        for date in sorted(df["_date"].dropna().unique()):
            day = df[df["_date"] == date].copy()

            daily_base = pd.DataFrame({"Name": daily_users})

            approved = day[
                (day[input_col].isin(daily_users)) &
                (day[status_col].astype(str).str.strip() == "1")
            ].copy()
            approved["_used"] = False

            # NOT FOUND
            not_found = day[
                (day[search_col].str.strip().str.lower() == "web") &
                (day[input_col].str.contains("nfuser", case=False, na=False)) &    
                (day[status_col].astype(str).str.strip() == "1")
            ]
            print(f"Date: {date} - Not Found Cases: {len(not_found)}")

            def match_daily_user(nfuser_name, daily_users):
                # Clean input: lower, remove 'nfuser', strip
                nfuser_clean = str(nfuser_name).lower().replace('nfuser', '').strip()

                # First try exact match
                for user in daily_users:
                    user_clean = user.lower()
                    for prefix in ["emp ", "int ", "freelancer "]:
                        if user_clean.startswith(prefix):
                            user_clean = user_clean[len(prefix):]
                    user_clean = user_clean.strip()
                    if nfuser_clean == user_clean:
                        return user

                # Fallback: if input is a unique prefix of a user name, allow match
                prefix_matches = []
                for user in daily_users:
                    user_clean = user.lower()
                    for prefix in ["emp ", "int ", "freelancer "]:
                        if user_clean.startswith(prefix):
                            user_clean = user_clean[len(prefix):]
                    user_clean = user_clean.strip()
                    if user_clean.startswith(nfuser_clean) and nfuser_clean:
                        prefix_matches.append(user)
                if len(prefix_matches) == 1:
                    return prefix_matches[0]
                return None

            not_found["_mapped_user"] = not_found[input_col].apply(
                lambda x: match_daily_user(x, daily_users)
            )

            # Count per daily_user
            not_found_counts = not_found.groupby("_mapped_user").size()

            # MULTIPLE CASES
            # dup_urls = approved[video_col].astype(str).str.strip()
            # dup_urls = dup_urls[dup_urls != ""]
            # dup_urls = dup_urls[dup_urls.duplicated(keep=False)]

            all_url_counts = day[website_col].astype(str).str.strip()
            all_url_counts = all_url_counts[all_url_counts != ""]
            all_url_counts = all_url_counts[all_url_counts.str.lower().isin(["nan", "none", "-"]) == False]
            url_counts = all_url_counts.value_counts()
            print(url_counts)
            dup_urls = url_counts[url_counts > 100].index

            multiple = approved[
                (~approved["_used"]) &
                (approved[input_col].isin(multi_users)) &
                (approved[search_col] == "Web") &
                (approved[wallet_col] == "UPI") &
                (approved[website_col].astype(str).str.strip().isin(dup_urls))
            ]
            approved.loc[multiple.index, "_used"] = True

            # DAILY CASES - First get all Web + UPI/Bank cases
            all_daily_candidates = approved[
                (~approved["_used"]) &
                (approved[search_col] == "Web") &
                (approved[wallet_col].isin(["UPI", "Bank Account"]))
            ]

            # Find URLs that appear more than 2 times - these go to Multiple Cases
            if not all_daily_candidates.empty:
                url_counts = all_daily_candidates[website_col].astype(str).str.strip()
                url_counts = url_counts[url_counts != ""]
                url_counts = url_counts[url_counts.str.lower().isin(["nan", "none", "-"]) == False]
                url_value_counts = url_counts.value_counts()
                # URLs appearing more than 2 times go to Multiple
                multiple_urls = url_value_counts[url_value_counts > 10].index

                # Separate into daily_cases (≤2 occurrences) and multiple (>2 occurrences)
                daily_cases = all_daily_candidates[
                    ~all_daily_candidates[website_col].astype(str).str.strip().isin(multiple_urls)
                ]
                # Additional multiple from daily candidates (URLs > 2 times)
                additional_multiple = all_daily_candidates[
                    all_daily_candidates[website_col].astype(str).str.strip().isin(multiple_urls)
                ]
            else:
                daily_cases = all_daily_candidates
                additional_multiple = approved[0:0]  # Empty DataFrame

            approved.loc[daily_cases.index, "_used"] = True

            # Add additional_multiple to multiple cases
            if not additional_multiple.empty:
                multiple = pd.concat([multiple, additional_multiple])
                approved.loc[additional_multiple.index, "_used"] = True

            # APP
            app_cases = approved[
                (~approved["_used"]) &
                (approved[search_col] == "App") &
                (approved[wallet_col] != "Crypto") &
                (approved[status_col].astype(str).str.strip() == "1")
            ]
            approved.loc[app_cases.index, "_used"] = True

            # CRYPTO
            crypto = approved[
                (~approved["_used"]) &
                (approved[search_col].isin(["Web", "App"])) &
                (approved[wallet_col] == "Crypto") &
                (approved[status_col].astype(str).str.strip() == "1")
            ]

            # WA/TG
            watg = approved[approved[search_col] == "Messaging Channel Platforms"]

            def m(d): return d.groupby(input_col).size()

            daily_base["Daily Cases"]    = daily_base["Name"].map(m(daily_cases)).fillna("NA")
            daily_base["Multiple Cases"] = daily_base["Name"].map(m(multiple)).fillna("NA")
            daily_base["Not Found"]      = daily_base["Name"].map(not_found_counts).fillna("NA")
            daily_base["App"]            = daily_base["Name"].map(m(app_cases)).fillna("NA")
            daily_base["WA/TG Case"]     = daily_base["Name"].map(m(watg)).fillna("NA")
            daily_base["Crypto Cases"]   = daily_base["Name"].map(m(crypto)).fillna("NA")
            daily_base["International Cases"] = "NA"  # populated from file 2

            # daily_base["Daily Cases"] = daily_base["Daily Cases"].replace(0, "NA")
            # daily_base["Multiple Cases"] = daily_base["Multiple Cases"].replace(0, "NA")

            num_cols_d = ["Daily Cases", "Multiple Cases", "Not Found", "App", "WA/TG Case", "Crypto Cases"]

            daily_base["Total Case"] = daily_base[num_cols_d].apply(
                lambda row: sum(
                    int(v) for v in row if str(v) != "NA" and str(v).replace(".", "", 1).isdigit()
                ), axis=1
            )

            daily_base["Total Case"] = daily_base.apply(
                lambda row: "NA" if all(str(row[c]) == "NA" for c in num_cols_d) else row["Total Case"],
                axis=1
            )

            # ERROR
            error_df = day[
                (day[status_col].astype(str).str.strip() == "2") &
                (day[input_col].astype(str).str.strip().str.lower() !=
                 day[qc_user_col].astype(str).str.strip().str.lower())
            ]
            # Map nfuser inputs to full names before grouping
            error_df = error_df.copy()
            error_df["_mapped_user"] = error_df[input_col].apply(
                lambda x: match_daily_user(x, daily_users)
            )
            # Use mapped user for grouping, but keep original if no mapping
            error_df["_group_col"] = error_df["_mapped_user"].fillna(error_df[input_col])
            daily_base["Error"] = daily_base["Name"].map(
                error_df.groupby("_group_col").size()
            ).fillna("NA")

            # QC SUMMARY
            emp_day = day[
                (day[qc_user_col].astype(str).str.strip().str.startswith("Emp")) &
                (day[input_col].astype(str).str.strip().str.lower() !=
                 day[qc_user_col].astype(str).str.strip().str.lower())
            ]

            total_qc = emp_day.groupby(qc_user_col).size()

            video_qc = emp_day[
                emp_day[video_col].notna() &
                (emp_day[video_col].astype(str).str.strip() != "") &
                (emp_day[video_col].astype(str).str.strip().str.lower() != "nan") &
                (emp_day[video_col].astype(str).str.strip().str.lower() != "none") &
                (emp_day[video_col].astype(str).str.strip().str.lower() != "-")
            ].groupby(qc_user_col).size()

            # ✅ Fix — reindex so all users appear even if video_qc = 0
            video_qc = video_qc.reindex(total_qc.index, fill_value=0)

            qc_data = pd.DataFrame({
                "Total QC": total_qc
            }).fillna(0).astype(int)

            # qc_data["Home QC"] = "NA"
            qc_data.reset_index(inplace=True)
            qc_data.rename(columns={qc_user_col: "Name"}, inplace=True)

            day_df = daily_base.merge(qc_data, on="Name", how="left")

            # For non-Emp users, set QC columns to NA
            qc_cols = ["Total QC"]
            for col in qc_cols:
                if col in day_df.columns:
                    day_df[col] = day_df.apply(
                        lambda r: r[col] if str(r["Name"]).startswith("Emp") else "NA", axis=1
                    )

            day_df = day_df.fillna("NA")

            # Convert floats to int
            for col in day_df.columns:
                if col == "Name":
                    continue
                day_df[col] = day_df[col].apply(
                    lambda x: int(float(x)) if x != "NA" and str(x).replace(".", "", 1).isdigit() else x
                )

            # TOTAL ROW
            total_row = {}
            for col in day_df.columns:
                if col == "Name":
                    total_row[col] = "Total"
                else:
                    col_data = day_df[col][day_df["Name"] != "Total"]  # ✅ existing total row exclude karo
                    numeric_vals = pd.to_numeric(col_data.replace("NA", None), errors="coerce")
                    total_row[col] = int(numeric_vals.sum()) if numeric_vals.notna().any() else "NA"

            day_df = pd.concat([day_df, pd.DataFrame([total_row])], ignore_index=True)
            day_df["Total Case"] = day_df["Total Case"].astype(object)
            daily_summary_all[date] = day_df

        df.drop(columns=["_date"], inplace=True, errors="ignore")

    daily_summary_df = pd.DataFrame()  # kept for Excel export compat

    summary_data = []
    multiple_summary = []        # FIX 1: collect across ALL dates (was only last date)
    freelancer_summary = []
    crypto_summary = []
    notFound_summary = []
    crypto_rows = {}

    with st.spinner("Processing data and checking database..."):
        for _, row in grouped.iterrows():
            date = row["Inserted_date"]
            cutoff_date = (pd.to_datetime(date) - timedelta(days=1)).strftime("%Y-%m-%d")

            # --- UPI new count ---
            date_upis = (
                upi_df.loc[upi_df["Inserted_date"] == date, "Upi_vpa_clean"]
                .dropna().astype(str).str.strip().str.lower().unique().tolist()
            )
            new_upi_today = count_new_upis_for_date(engine, date_upis, cutoff_date)

            # --- Bank new count ---
            date_banks = (
                bank_df.loc[bank_df["Inserted_date"] == date, "Bank_acc_clean"]
                .dropna().astype(str).str.strip().unique().tolist()
            )
            new_bank_today = count_new_banks_for_date(engine, date_banks, cutoff_date)

            total_upi = int(row["Total_UPI"]) if not pd.isna(row["Total_UPI"]) else 0
            unique_upi = int(row["Unique_UPI"]) if not pd.isna(row["Unique_UPI"]) else 0
            bank_total = int(row["Bank_Total"]) if not pd.isna(row["Bank_Total"]) else 0
            bank_unique = int(row["Bank_Unique"]) if not pd.isna(row["Bank_Unique"]) else 0

            summary_data.append({
                "Date": date,
                "Total": int(row["website_total"]),
                "UPI_Total": total_upi,
                "UPI_Unique": unique_upi,
                "UPI_pct": f"{(unique_upi / total_upi * 100):.0f}%" if total_upi else "0%",
                "UPI_New": new_upi_today,
                "UPI_New_pct": f"{(new_upi_today / unique_upi * 100):.0f}%" if unique_upi else "0%",
                "Bank_Total": bank_total,
                "Bank_Unique": bank_unique,
                "Bank_pct": f"{(bank_unique / bank_total * 100):.0f}%" if bank_total else "0%",
                "Bank_New": new_bank_today,
                "Bank_New_pct": f"{(new_bank_today / bank_unique * 100):.0f}%" if bank_unique else "0%",
                "unique_website": int(row["unique_website"]) if not pd.isna(row["unique_website"]) else 0
            })

            # FIX 1: Multiple user summary — now runs for EVERY date
            for user in target_users:
                duplicate_urls = (
                    upi_df[upi_df["Inserted_date"] == date]
                    .groupby(["Input_user", "Website_url"])["Website_url"]
                    .transform("count") > 50
                )

                user_mask = (
                    (upi_df["Inserted_date"] == date) &
                    (upi_df["Input_user"].astype(str).str.strip() == user) & duplicate_urls
                )
                user_sub = upi_df.loc[user_mask].copy()
                total = int(len(user_sub))
                unique_count = int(user_sub["Upi_vpa_clean"].dropna().astype(str).str.strip().nunique())
                user_upis_list = user_sub["Upi_vpa_clean"].dropna().astype(str).str.strip().unique().tolist()
                new_count = count_new_upis_for_date(engine, user_upis_list, cutoff_date) if user_upis_list else 0

                web_count = int(
                    user_sub["Search_for"].astype(str).str.strip().str.lower()
                    .eq("web").sum()
                ) if "Search_for" in user_sub.columns else 0

                app_count = int(
                    user_sub["Search_for"].astype(str).str.strip().str.lower()
                    .eq("app").sum()
                ) if "Search_for" in user_sub.columns else 0
                print(app_count)

                if total == 0:
                    multiple_summary.append({
                        "Date": date,
                        "Input_user": user,
                        "Total": "AB",
                        "Web_Count": "AB",
                        "App_Count": "AB",
                        "Unique_UPI_Count": "AB",
                        "Unique_UPI_pct": "AB",
                        "New_UPI_Count": "AB",
                        "New_UPI_pct": "AB"
                    })
                else:
                    multiple_summary.append({
                        "Date": date,
                        "Input_user": user,
                        "Total": total,
                        "Web_Count": web_count,
                        "App_Count": app_count,
                        "Unique_UPI_Count": unique_count,
                        "Unique_UPI_pct": f"{(unique_count / total * 100):.0f}%" if total else "0%",
                        "New_UPI_Count": new_count,
                        "New_UPI_pct": f"{(new_count / unique_count * 100):.0f}%" if unique_count else "0%"
                    })

        # FIX 2: Freelancer loop — was correct structure but reuse same cutoff_date
        for date in sorted(all_dates):
            cutoff_date = (pd.to_datetime(date) - timedelta(days=1)).strftime("%Y-%m-%d")

            date_series = df["Inserted_date"].pipe(pd.to_datetime, errors="coerce").dt.date

            freelancer_df = df.loc[
                (date_series == date) &
                (df["Input_user"].astype(str).str.contains("Freelancer", case=False, na=False)) &
                (df["Approvd_status"].astype(str).str.strip() == "1")
            ].copy()

            int_df = df.loc[
                (date_series == date) &
                (df["Input_user"].astype(str).str.contains("INT", case=False, na=False)) &
                (~df["Input_user"].astype(str).str.contains("icuser", case=False, na=False)) &
                (df["Approvd_status"].astype(str).str.strip() == "1")
            ].copy()

            emp_df = df.loc[
                (date_series == date) &
                (df["Input_user"].astype(str).str.contains("Emp", case=False, na=False)) &
                (~df["Input_user"].astype(str).str.contains("icuser", case=False, na=False)) &
                (df["Approvd_status"].astype(str).str.strip() == "1")
            ].copy()

            for user_type, sub_df in [("Freelancer", freelancer_df), ("INT", int_df), ("Employee", emp_df)]:
                t_upi, u_upi, n_upi, t_bank, u_bank, n_bank = process_df(sub_df, engine, cutoff_date)
                freelancer_summary.append({
                    "User_Type": user_type,
                    "Date": date,
                    "Total_UPI": t_upi,
                    "Unique_UPI": u_upi,
                    "New_UPI": n_upi,
                    "Total_Bank": t_bank,
                    "Unique_Bank": u_bank,
                    "New_Bank": n_bank,
                    "Total Cases": t_upi + t_bank
                })

        for date in sorted(all_dates):
            crypto_filtered = df.loc[
                (df["Inserted_date"].pipe(pd.to_datetime, errors="coerce").dt.date == date) &
                (df["Feature_type"].astype(str).str.strip() == "BS Money Laundering") &
                (df["Upi_bank_account_wallet"].astype(str).str.strip().str.lower() == "crypto") &
                (df["Approvd_status"].astype(str).str.strip() == "1")
            ].copy()

            if crypto_filtered.empty:
                continue

            for user, user_df in crypto_filtered.groupby("Input_user"):
                # website_count = int(
                #     user_df["Website_url"].dropna().astype(str).str.strip().nunique()
                # ) if "Website_url" in user_df.columns else 0

                crypto_app_count = int(user_df[
                    user_df["Search_for"].astype(str).str.strip().str.lower() == "app"
                ].shape[0]) if "Search_for" in user_df.columns else 0

                crypto_web_count = int(user_df[
                    user_df["Search_for"].astype(str).str.strip().str.lower() == "web"
                ].shape[0]) if "Search_for" in user_df.columns else 0

                crypto_summary.append({
                    "Date": date,
                    "Name": str(user).strip(),
                    "Crypto App": crypto_app_count,
                    "Crypto Web": crypto_web_count,
                    "Total": crypto_app_count + crypto_web_count
                })
        
        for date in sorted(all_dates):
            notFound_Filtered = df.loc[
                (df["Inserted_date"].pipe(pd.to_datetime, errors="coerce").dt.date == date) &
                (df["Feature_type"].astype(str).str.strip() == "BS Money Laundering") &
                (df["Search_for"].astype(str).str.strip() == "Web") &
                (df["Input_user"].astype(str).str.contains("nfuser", case=False, na=False))
            ].copy()

            if notFound_Filtered.empty:
                continue

            for user, user_df in notFound_Filtered.groupby("Input_user"):
                approved_count = int(user_df[
                    user_df["Approvd_status"].astype(str).str.strip() == "1"
                ].shape[0]) if "Approvd_status" in user_df.columns else 0

                rejected_count = int(user_df[
                    (user_df["Approvd_status"].astype(str).str.strip() == "2") &
                    (user_df["Input_user"].astype(str).str.strip().str.lower() != user_df["Approved_by"].astype(str).str.strip().str.lower())
                ].shape[0]) if "Approvd_status" in user_df.columns else 0   

                notFound_summary.append({
                    "Date": date,
                    "Name": str(user).strip(),
                    "Approved": approved_count,
                    "Rejected": rejected_count
                })

    summary_df = pd.DataFrame(summary_data)
    multiple_summary_df = pd.DataFrame(multiple_summary)
    freelancer_summary_df = pd.DataFrame(freelancer_summary)
    crypto_summary_df = pd.DataFrame(crypto_summary)
    notFound_summary_df = pd.DataFrame(notFound_summary)

    # ================= DISPLAY =================
    st.subheader("📊 Summary Report")

    summary_col, date_col = st.columns(2)

    # ========== DROPDOWNS — Side by Side ==========
    col1, col2 = st.columns(2)

    with col1:
        summary_type = st.selectbox(
            "Select Summary Type",
            ["UPI & Bank Summary", "Multiple User's Summary", "Crypto Summary",
            "Employee, Intern & Freelancer Summary", "Daily Summary", "Not Found Summary"],
            key="summary_select"
        )

    with col2:
        selected_date = None
        if summary_type == "Multiple User's Summary" and not multiple_summary_df.empty:
            available_dates = sorted(multiple_summary_df["Date"].unique())
            selected_date = st.selectbox("Select Date", available_dates, index=len(available_dates) - 1)

        elif summary_type == "Daily Summary" and daily_summary_all:
            available_dates = sorted(daily_summary_all.keys())
            selected_date = st.selectbox("Select Date", available_dates, index=len(available_dates) - 1)

        elif summary_type == "Crypto Summary" and not crypto_summary_df.empty:
            available_dates = sorted(crypto_summary_df["Date"].unique())
            selected_date = st.selectbox("Select Date", available_dates, index=len(available_dates) - 1)

        elif summary_type == "Not Found Summary" and not notFound_summary_df.empty:
            available_dates = sorted(notFound_summary_df["Date"].unique())
            selected_date = st.selectbox("Select Date", available_dates, index=len(available_dates) - 1)

        elif summary_type == "Employee, Intern & Freelancer Summary" and not freelancer_summary_df.empty:
            available_dates = sorted(freelancer_summary_df["Date"].unique())
            selected_date = st.selectbox("Select Date", available_dates, index=len(available_dates) - 1)

    # ========== TABLES — Full Width ==========
    website_searching   = {}
    credential_making   = {}
    remark_checking     = {}
    upi_fraud           = {}
    remark_col          = {}
    international_cases = {}

    if summary_type == "UPI & Bank Summary":
        html_table = """
        <style>
        .table-container { width:100%; overflow-x:auto; }
        .excel-table { border-collapse:collapse; font-family:'Segoe UI',sans-serif; font-size:13px; width:100% !important; table-layout:fixed !important; }
        .excel-table th, .excel-table td { border:1px solid #ccc; text-align:center; padding:6px 4px; white-space:normal; word-wrap:break-word; }
        .excel-table thead tr:first-child th { background:#cbd5e1; font-size:16px; font-weight:700; }
        .excel-table thead tr:nth-child(2) th { background:#cbd5e1; font-size:14px; font-weight:600; }
        .excel-table thead tr:nth-child(3) th { background:#e2e8f0; font-size:12px; }
        .excel-table td { background:#f8fafc; }
        </style>
        <div class="table-container"><table class="excel-table">
        <thead>
            <tr><th colspan="13">UPI, Bank & Website Report</th></tr>
            <tr>
                <th rowspan="2">Date</th><th rowspan="2">Total</th>
                <th colspan="5">UPI</th><th colspan="5">Bank</th>
                <th rowspan="2">Unique Website</th>
            </tr>
            <tr>
                <th>Total</th><th>Unique</th><th>%</th><th>New</th><th>%</th>
                <th>Total</th><th>Unique</th><th>%</th><th>New</th><th>%</th>
            </tr>
        </thead><tbody>
        """
        for _, row in summary_df.iterrows():
            html_table += f"""<tr>
                <td>{row['Date']}</td><td>{row['Total']}</td>
                <td>{row['UPI_Total']}</td><td>{row['UPI_Unique']}</td><td>{row['UPI_pct']}</td><td>{row['UPI_New']}</td><td>{row['UPI_New_pct']}</td>
                <td>{row['Bank_Total']}</td><td>{row['Bank_Unique']}</td><td>{row['Bank_pct']}</td><td>{row['Bank_New']}</td><td>{row['Bank_New_pct']}</td>
                <td>{row['unique_website']}</td>
            </tr>"""
        html_table += "</tbody></table></div>"
        wrapped, h = with_copy_button(html_table, table_id="upi-bank-table", height=450)
        components.html(wrapped, height=h, scrolling=True)

    elif summary_type == "Multiple User's Summary":
        if selected_date and not multiple_summary_df.empty:
            day_df = multiple_summary_df[multiple_summary_df["Date"] == selected_date]

            num_df = day_df[day_df["Total"] != "AB"]
            total_total = pd.to_numeric(num_df["Total"], errors="coerce").sum()
            total_unique = pd.to_numeric(num_df["Unique_UPI_Count"], errors="coerce").sum()
            total_new = pd.to_numeric(num_df["New_UPI_Count"], errors="coerce").sum()
            total_unique_pct = f"{(total_unique / total_total * 100):.0f}%" if total_total else "0%"
            total_new_pct = f"{(total_new / total_unique * 100):.0f}%" if total_unique else "0%"
            total_web = pd.to_numeric(num_df["Web_Count"], errors="coerce").sum()
            total_app = pd.to_numeric(num_df["App_Count"], errors="coerce").sum()

            table = f"""
            <style>
            .table-user {{ width:100%; border-collapse:collapse; font-family:'Segoe UI',sans-serif; font-size:14px; }}
            .table-user th, .table-user td {{ border:1px solid #000; padding:6px 10px; text-align:center; }}
            .table-user thead th {{ background:#cfe8b0; font-weight:700; }}
            .table-user tfoot td {{ font-weight:700; background:#cfe8b0; }}
            .ab-cell {{ background:#e9dcc2 !important; font-weight:700; }}
            </style>
            <table class="table-user">
            <thead>
                <tr><th colspan="8">Multiple User's Counts ({selected_date})</th></tr>
                <tr>
                    <th rowspan="2">Name</th>
                    <th rowspan="2">Total</th>
                    <th rowspan="2">Web Count</th>
                    <th rowspan="2">App Count</th>
                    <th colspan="2">Unique UPI</th>
                    <th colspan="2">New UPI</th>
                </tr>
                <tr>
                    <th>Count</th><th>%</th>
                    <th>Count</th><th>%</th>
                </tr>
            </thead>
            <tbody>
            """
            for _, row in day_df.iterrows():
                is_ab = str(row['Total']).strip() == 'AB'
                row_style = " style='background:#e9dcc2;'" if is_ab else ""
                ab_css = ' class="ab-cell"' if is_ab else ""
                table += f"""<tr{row_style}>
                    <td style="text-align:left">{row['Input_user']}</td>
                    <td{ab_css}>{row['Total']}</td>
                    <td{ab_css}>{row['Web_Count']}</td>
                    <td{ab_css}>{row['App_Count']}</td>
                    <td{ab_css}>{row['Unique_UPI_Count']}</td>
                    <td{ab_css}>{row['Unique_UPI_pct']}</td>
                    <td{ab_css}>{row['New_UPI_Count']}</td>
                    <td{ab_css}>{row['New_UPI_pct']}</td>
                </tr>"""

            table += f"""</tbody>
            <tfoot>
                <tr>
                    <td style="text-align:left">Total</td>
                    <td>{total_total}</td>
                    <td>{total_web}</td>
                    <td>{total_app}</td>
                    <td>{total_unique}</td>
                    <td>{total_unique_pct}</td>
                    <td>{total_new}</td>
                    <td>{total_new_pct}</td>
                </tr>
            </tfoot>
            </table>"""
            wrapped, h = with_copy_button(table, table_id="multiple-user-table", height=450)
            components.html(wrapped, height=h, scrolling=True)

    elif summary_type == "Employee, Intern & Freelancer Summary":
        if selected_date and not freelancer_summary_df.empty:
            day_df = freelancer_summary_df[freelancer_summary_df["Date"] == selected_date]

            def get_row(user_type):
                r = day_df[day_df["User_Type"] == user_type]
                if r.empty:
                    return [0] * 7
                r = r.iloc[0]
                return [r["Total_UPI"], r["Unique_UPI"], r["New_UPI"], r["Total_Bank"], r["Unique_Bank"], r["New_Bank"], r["Total Cases"]]

            e = get_row("Employee")
            i = get_row("INT")
            f = get_row("Freelancer")
            totals = [e[j] + i[j] + f[j] for j in range(7)]

            table = f"""
            <style>
            .table-user {{ width:100%; border-collapse:collapse; font-family:'Segoe UI',sans-serif; font-size:14px; }}
            .table-user th, .table-user td {{ border:1px solid #000; padding:6px 10px; text-align:center; }}
            .table-user thead th {{ background:#cfe8b0; font-weight:700; }}
            .table-user tfoot td {{ font-weight:700; }}
            </style>
            <table class="table-user" id="summary-table-3">
            <thead>
                <tr><th colspan="9">Employee, Intern & Freelancer Summary ({selected_date})</th></tr>
                <tr>
                    <th rowspan="2">User</th>
                    <th rowspan="2">Date</th>
                    <th colspan="3">UPI</th>
                    <th colspan="3">Bank</th>
                    <th rowspan="2">Total Cases</th>
                </tr>
                <tr><th>Total</th><th>Unique</th><th>New</th><th>Total</th><th>Unique</th><th>New</th></tr>
            </thead><tbody>
                <tr><td>Employee</td><td rowspan="3">{selected_date}</td><td>{e[0]}</td><td>{e[1]}</td><td>{e[2]}</td><td>{e[3]}</td><td>{e[4]}</td><td>{e[5]}</td><td>{e[6]}</td></tr>
                <tr><td>Intern</td><td>{i[0]}</td><td>{i[1]}</td><td>{i[2]}</td><td>{i[3]}</td><td>{i[4]}</td><td>{i[5]}</td><td>{i[6]}</td></tr>
                <tr><td>Freelancer</td><td>{f[0]}</td><td>{f[1]}</td><td>{f[2]}</td><td>{f[3]}</td><td>{f[4]}</td><td>{f[5]}</td><td>{f[6]}</td></tr>
            </tbody>
            <tfoot><tr>
                <td>Total</td><td>NA</td>
                <td>{totals[0]}</td><td>{totals[1]}</td><td>{totals[2]}</td>
                <td>{totals[3]}</td><td>{totals[4]}</td><td>{totals[5]}</td>
                <td>{totals[6]}</td>
            </tr></tfoot></table>"""
            wrapped, h = with_copy_button(table, table_id="emp-intern-freelancer-table", height=450)
            components.html(wrapped, height=h, scrolling=True)

    elif summary_type == "Crypto Summary":
        if selected_date and not crypto_summary_df.empty:
            day_df = crypto_summary_df[crypto_summary_df["Date"] == selected_date]

            display_crypto = (
                day_df.groupby("Name")[["Crypto App", "Crypto Web", "Total"]]
                .sum()
                .reset_index()
            )

            total_count = display_crypto["Total"].sum()
            total_app     = display_crypto["Crypto App"].sum()
            total_web     = display_crypto["Crypto Web"].sum()

            table = f"""
            <style>
            .table-crypto {{ width:100%; border-collapse:collapse; font-family:'Segoe UI',sans-serif; font-size:14px; }}
            .table-crypto th, .table-crypto td {{ border:1px solid #000; padding:6px 10px; text-align:center; }}
            .table-crypto thead th {{ background:#ebb7ac; font-weight:700; }}
            .table-crypto tfoot td {{ font-weight:700; background:#ebb7ac; }}
            </style>
            <table class="table-crypto">
            <thead>
                <tr><th colspan="4">Crypto User's Counts ({selected_date})</th></tr>
                <tr>
                    <th>Name</th>
                    <th>Crypto App</th>
                    <th>Crypto Web</th>
                    <th>Total</th>
                </tr>
            </thead>
            <tbody>
            """
            for _, row in display_crypto.iterrows():
                table += f"""<tr>
                    <td style="text-align:left">{row['Name']}</td>
                    <td>{row['Crypto App']}</td>
                    <td>{row['Crypto Web']}</td>
                    <td>{row['Total']}</td>
                </tr>"""

            table += f"""</tbody>
            <tfoot>
                <tr>
                    <td style="text-align:left">Total</td>
                    <td>{total_app}</td>
                    <td>{total_web}</td>
                    <td>{total_count}</td>
                </tr>
            </tfoot>
            </table>"""
            wrapped, h = with_copy_button(table, table_id="crypto-table", height=400)
            components.html(wrapped, height=h, scrolling=True)

    elif summary_type == "Not Found Summary":
        if selected_date and not notFound_summary_df.empty:
            day_df = notFound_summary_df[notFound_summary_df["Date"] == selected_date]

            display_df = (
                day_df.groupby("Name")[["Approved", "Rejected"]]
                .sum()
                .reset_index()
            )

            total_approved = display_df["Approved"].sum()
            total_rejected = display_df["Rejected"].sum()

            table = f"""
            <style>
            .table-notfound {{ width:40%; border-collapse:collapse; font-family:'Segoe UI',sans-serif; font-size:14px; }}
            .table-notfound th, .table-notfound td {{ border:1px solid #000; padding:6px 10px; text-align:center; }}
            .table-notfound thead th {{ background:#ffffff; font-weight:700; }}
            .table-notfound .footer-approved td {{ font-weight:700; }}
            </style>
            <table class="table-notfound">
            <thead>
                <tr><th colspan="3">{selected_date}</th></tr>
                <tr>
                    <th style="text-align:left">Input_user</th>
                    <th>Approved</th>
                    <th>Rejected</th>
                </tr>
            </thead>
            <tbody>
            """
            for _, row in display_df.iterrows():
                table += f"""<tr>
                    <td style="text-align:left">{row['Name']}</td>
                    <td>{row['Approved']}</td>
                    <td>{row['Rejected']}</td>
                </tr>"""

            table += f"""</tbody>
            <tfoot>
                <tr class="footer-approved">
                    <td style="text-align:left"><b>Total</b></td>
                    <td style="background:#ffff00;"><b>{total_approved}</b></td>
                    <td><b>{total_rejected}</b></td>
                </tr>
            </tfoot>
            </table>"""
            wrapped, h = with_copy_button(table, table_id="notfound-table", height=450)
            components.html(wrapped, height=h, scrolling=True)


    elif summary_type == "Daily Summary":
        if daily_summary_all and selected_date:
            day_df = daily_summary_all[selected_date].copy()

            col_mapping = {
                "Daily Cases":                "Daily Cases",
                "Multiple Cases":             "Multiple Cases",
                "Not Found Cases":            "Not Found",
                "App":                        "App",
                "Messaging Channel Platform": "WA/TG Case",
                "Crypto cases":                "Crypto Cases",
                "Errors":                     "Error",
            }

            mismatches = {}
            
            if uploaded_file_2 is not None:
                try:
                    df2 = pd.read_excel(uploaded_file_2)
                    df2.columns = df2.columns.str.strip()

                    date_col_2 = None
                    for c in df2.columns:
                        if c.strip().lower() in ["date", "inserted date", "inserted_date"]:
                            date_col_2 = c
                            break

                    if date_col_2:
                        df2[date_col_2] = pd.to_datetime(df2[date_col_2], errors="coerce").dt.date
                        df2_day = df2[df2[date_col_2] == selected_date].copy()
                    else:
                        df2_day = df2.copy()

                    df2_day["user name"] = df2_day["user name"].astype(str).str.strip().str.lower()
                    df2_grouped = df2_day.groupby("user name")[list(col_mapping.keys())].sum().reset_index()

                    def safe_int_val(val):
                        if pd.isna(val):
                            return 0
                        s = str(val).strip().replace(".", "", 1)
                        return int(float(str(val))) if s.isdigit() else 0

                    for _, r in df2_day.iterrows():
                        u = str(r["user name"]).strip().lower()
                        website_searching[u] = website_searching.get(u, 0) + safe_int_val(r.get("Website Searching", 0))
                        credential_making[u] = credential_making.get(u, 0) + safe_int_val(r.get("Credential Making", 0))
                        remark_checking[u]   = remark_checking.get(u, 0)   + safe_int_val(r.get("Remark Checking", 0))
                        upi_fraud[u]         = upi_fraud.get(u, 0)         + safe_int_val(r.get("UPI Fraud", 0))
                        international_cases[u] = international_cases.get(u, 0) + safe_int_val(r.get("International cases", 0))

                        new_val  = str(r.get("remark", "")).strip()
                        existing = remark_col.get(u, "")
                        if new_val and new_val.lower() not in ["nan", "none", ""]:
                            remark_col[u] = (existing + " | " + new_val).strip(" | ") if existing else new_val
                        else:
                            remark_col[u] = existing if existing else "NA"

                    exclude_highlight_users = ["Emp Shubhankar Shukla", "Emp Sheetal Dubey"]

                    def row_is_all_na(row, cols):
                        return all(str(row[c]).strip() in ("NA", "") for c in cols)

                    for idx, row in day_df.iterrows():
                        name = str(row["Name"]).strip()
                        if name == "Total" or name in exclude_highlight_users:
                            continue

                        clean_name = name.strip()
                        for prefix in ["Emp ", "INT ", "emp ", "int "]:
                            if clean_name.startswith(prefix):
                                clean_name = clean_name[len(prefix):]
                                break

                        name_lower = clean_name.strip().lower()
                        match = df2_grouped[df2_grouped["user name"] == name_lower]

                        if match.empty:
                            continue

                        num_cols = ["Daily Cases", "Multiple Cases", "Not Found", "App", "WA/TG Case", "Crypto Cases", "Error"]
                        if row_is_all_na(row, [c for c in num_cols if c in day_df.columns]):
                            for col2, col1 in col_mapping.items():
                                if col1 not in day_df.columns:
                                    continue
                                val2 = match.iloc[0][col2]
                                if pd.isna(val2):
                                    continue
                                s2 = str(val2).strip()
                                if s2.lower() in ("", "na", "nan", "none"):
                                    continue
                                if s2.replace('.', '', 1).isdigit():
                                    day_df.at[idx, col1] = int(float(s2))
                                else:
                                    day_df.at[idx, col1] = s2

                            # Recalculate Total Case for this row after filling second-file values
                            total_val = 0
                            for col in ["Daily Cases", "Multiple Cases", "Not Found", "App", "WA/TG Case", "Crypto Cases"]:
                                if col in day_df.columns:
                                    v = day_df.at[idx, col]
                                    if str(v).replace('.', '', 1).isdigit():
                                        total_val += int(float(str(v)))
                            day_df.at[idx, "Total Case"] = total_val if total_val != 0 else "NA"

                        def normalize_compare_value(value):
                            if pd.isna(value):
                                return 0
                            if isinstance(value, str):
                                s = value.strip()
                                if s.lower() in ("", "na", "nan", "none"):
                                    return 0
                                if s == "AB":
                                    return 0
                                if s.replace('.', '', 1).isdigit():
                                    return int(float(s))
                                return s
                            try:
                                return int(float(value))
                            except Exception:
                                return value

                        for col2, col1 in col_mapping.items():
                            if col1 not in day_df.columns:
                                continue
                            val1 = row[col1]
                            val2 = match.iloc[0][col2]
                            v1 = normalize_compare_value(val1)
                            v2 = normalize_compare_value(val2)

                            if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
                                if int(v1) != int(v2):
                                    mismatches[(name, col1)] = int(v2)
                            elif v1 != v2:
                                mismatches[(name, col1)] = v2

                except Exception as e:
                    st.warning(f"Second file read error: {e}")

            def get_val(d, name):
                name_lower = name.strip().lower()
                if name_lower in d:
                    return d[name_lower]
                clean = name.strip()
                for prefix in ["Emp ", "INT ", "emp ", "int "]:
                    if clean.startswith(prefix):
                        clean = clean[len(prefix):]
                        break
                clean_lower = clean.strip().lower()
                if clean_lower in d:
                    return d[clean_lower]
                for key in d:
                    if clean_lower in key or key in clean_lower:
                        return d[key]
                return "NA"

            def has_second_sheet_presence(name):
                if not name or name == "Total":
                    return False
                for d in [website_searching, credential_making, remark_checking, upi_fraud, international_cases, remark_col]:
                    val = get_val(d, name)
                    if val != "NA":
                        return True
                return False

            # ── Merge International Cases from file 2 into day_df ──────────
            if uploaded_file_2 is not None:
                for idx, row in day_df.iterrows():
                    name = str(row["Name"]).strip()
                    if name == "Total":
                        continue
                    val = get_val(international_cases, name)
                    new_ic = str(val) if (val not in ("NA", 0)) else "NA"
                    day_df.at[idx, "International Cases"] = new_ic

                    # ── Add International Cases to Total Case ──────────────
                    if new_ic not in ("NA", "AB"):
                        old_tc = day_df.at[idx, "Total Case"]
                        if str(old_tc) not in ("NA", "AB"):
                            day_df.at[idx, "Total Case"] = int(float(str(old_tc))) + int(new_ic)
                        else:
                            day_df.at[idx, "Total Case"] = str(int(new_ic))

                # Recompute Total row for International Cases and Total Case
                non_total = day_df[day_df["Name"] != "Total"]
                total_idx = day_df[day_df["Name"] == "Total"].index

                ic_vals = pd.to_numeric(non_total["International Cases"].replace({"NA": None, "AB": None}), errors="coerce")
                total_ic = int(ic_vals.sum()) if ic_vals.notna().any() else "NA"

                tc_vals = pd.to_numeric(non_total["Total Case"].replace({"NA": None, "AB": None}), errors="coerce")
                total_tc = int(tc_vals.sum()) if tc_vals.notna().any() else "NA"

                if len(total_idx) > 0:
                    day_df.at[total_idx[0], "International Cases"] = str(total_ic)
                    day_df.at[total_idx[0], "Total Case"]          = str(total_tc)

                # ── Recompute AB rows after second sheet data is applied ──────────
                data_cols_ab = [c for c in day_df.columns if c != "Name"]
                for idx, row in day_df.iterrows():
                    if str(row["Name"]).strip() == "Total":
                        continue
                    name = str(row["Name"]).strip()
                    if has_second_sheet_presence(name):
                        for col in data_cols_ab:
                            if str(day_df.at[idx, col]).strip() == "AB":
                                day_df.at[idx, col] = "NA"
                        continue
                    if all(str(row[col]).strip() in ("NA", "") for col in data_cols_ab):
                        for col in data_cols_ab:
                            day_df.at[idx, col] = "AB"
                    elif str(row.get("Daily Cases", "")).strip() == "AB":
                        for col in data_cols_ab:
                            day_df.at[idx, col] = "AB"

                if "Daily Cases" in day_df.columns:
                    day_df["Daily Cases"] = day_df["Daily Cases"].replace(0, "NA")
                if "Multiple Cases" in day_df.columns:
                    day_df["Multiple Cases"] = day_df["Multiple Cases"].replace(0, "NA")
                if "Not Found" in day_df.columns:
                    day_df["Not Found"] = day_df["Not Found"].replace(0, "NA")
                if "App" in day_df.columns:
                    day_df["App"] = day_df["App"].replace(0, "NA")
                if "WA/TG Case" in day_df.columns:
                    day_df["WA/TG Case"] = day_df["WA/TG Case"].replace(0, "NA")
                if "Crypto Cases" in day_df.columns:
                    day_df["Crypto Cases"] = day_df["Crypto Cases"].replace(0, "NA")
                if "International Cases" in day_df.columns:
                    day_df["International Cases"] = day_df["International Cases"].replace(0, "NA")
                if "Error" in day_df.columns:
                    day_df["Error"] = day_df["Error"].replace(0, "NA")

                # ── Sync back so Excel export uses updated values ──────────
                daily_summary_all[selected_date] = day_df.copy()

            base_colspan  = 11
            extra_colspan = 5 if uploaded_file_2 is not None else 0
            total_colspan = base_colspan + extra_colspan

            extra_section_headers = ""
            extra_col_headers     = ""
            if uploaded_file_2 is not None:
                extra_section_headers = """
                    <th colspan="3" style="background:#d9e1f2; font-weight:700;">Website Monitoring</th>
                    <th colspan="1" style="background:#d9e1f2; font-weight:700;">SM Scrawling</th>
                    <th colspan="1" style="background:#d9e1f2; font-weight:700;">Remark</th>
                """
                extra_col_headers = """
                    <th style="background:#d9e1f2;">Website<br>Searching</th>
                    <th style="background:#d9e1f2;">Remarks<br>Checking</th>
                    <th style="background:#d9e1f2;">Credentials</th>
                    <th style="background:#d9e1f2;">UPI Fraud</th>
                    <th style="background:#d9e1f2;">Remark</th>
                """

            columns_order = ["Name", "Daily Cases", "Multiple Cases", "Not Found", "App",
                            "WA/TG Case", "Crypto Cases", "International Cases", "Total Case", "Error",
                            "Total QC"]
            columns_order = [c for c in columns_order if c in day_df.columns]

            shared_css = """
            <style>
            .table-daily, .table-daily2 {
                width:100%; border-collapse:collapse;
                font-family:'Segoe UI',sans-serif; font-size:13px;
            }
            .table-daily th, .table-daily td,
            .table-daily2 th, .table-daily2 td {
                border:1px solid #000; padding:5px 5px;
                text-align:center; white-space:nowrap;
            }
            .mismatch { background:#ff4d4d !important; color:#fff; font-weight:700; }
            .ab-cell  { background:#e9dcc2 !important; font-weight:700; }
            .ab-cell2 { background:#e9dcc2 !important; font-weight:700; }
            </style>
            """

            def build_header(table_class):
                return f"""
                <table class="{table_class}">
                <thead>
                    <tr>
                        <th colspan="{total_colspan}" style="background:#b8cce4; font-size:15px; text-align:center; font-weight:700;">
                            Daily wise user summary
                        </th>
                    </tr>
                    <tr>
                        <th colspan="{total_colspan}" style="background:#dce6f1; font-size:14px; text-align:center; font-weight:700;">
                            {selected_date}
                        </th>
                    </tr>
                    <tr>
                        <th rowspan="2" style="background:#d9e1f2; font-weight:700; width:10%;">Name</th>
                        <th colspan="9" style="background:#d9e1f2; font-weight:700;">Insertion</th>
                        <th colspan="1" style="background:#dce6f1; font-weight:700;">Quality Check</th>
                        {extra_section_headers}
                    </tr>
                    <tr>
                        <th style="background:#d9e1f2;">Daily<br>Cases</th>
                        <th style="background:#d9e1f2;">Multiple<br>Cases</th>
                        <th style="background:#d9e1f2;">Not<br>Found</th>
                        <th style="background:#d9e1f2;">App</th>
                        <th style="background:#d9e1f2;">WA/TG<br>Case</th>
                        <th style="background:#d9e1f2;">Crypto<br>Cases</th>
                        <th style="background:#d9e1f2;">International<br>Cases</th>
                        <th style="background:#d9e1f2;">Total<br>Case</th>
                        <th style="background:#d9e1f2;">Error</th>
                        <th style="background:#dce6f2;">Total QC</th>
                        {extra_col_headers}
                    </tr>
                </thead>
                <tbody>
                """

            # ===== TABLE 1 — WITH COMPARISON =====
            table1 = build_header("table-daily")
            total_row1 = None
            ws_total1 = cr_total1 = rc_total1 = uf_total1 = 0

            for _, row in day_df.iterrows():
                name = str(row["Name"]).strip()
                if name == "Total":
                    total_row1 = row
                    continue

                is_ab = str(row.get("Daily Cases", "")).strip() == "AB"
                row_style = " style='background:#e9dcc2;'" if is_ab else ""
                table1 += f"<tr{row_style}>"

                for col in columns_order:
                    val = row[col] if col in row else "NA"
                    is_mismatch = (name, col) in mismatches

                    if is_ab and col != "Name":
                        css = ' class="ab-cell"'
                        display_val = "AB"
                    elif is_mismatch:
                        css = ' class="mismatch"'
                        display_val = mismatches[(name, col)]
                    elif str(val) == "AB":
                        css = ' class="ab-cell"'
                        display_val = val
                    else:
                        css = ''
                        display_val = val

                    align = ' style="text-align:left"' if col == "Name" else ''
                    table1 += f'<td{css}{align}>{display_val}</td>'

                if uploaded_file_2 is not None:
                    if is_ab:
                        ws = rc = cr = uf = rem = "AB"
                    else:
                        ws  = get_val(website_searching, name)
                        rc  = get_val(remark_checking, name)
                        cr  = get_val(credential_making, name)
                        uf  = get_val(upi_fraud, name)
                        rem = get_val(remark_col, name)

                        ws = "NA" if ws == 0 or ws == "NA" else ws
                        rc = "NA" if rc == 0 or rc == "NA" else rc
                        cr = "NA" if cr == 0 or cr == "NA" else cr
                        uf = "NA" if uf == 0 or uf == "NA" else uf

                    if str(ws) not in ("NA", "AB"): ws_total1 += int(ws)
                    if str(rc) not in ("NA", "AB"): rc_total1 += int(rc)
                    if str(cr) not in ("NA", "AB"): cr_total1 += int(cr)
                    if str(uf) not in ("NA", "AB"): uf_total1 += int(uf)

                    ab_css = ' class="ab-cell"' if is_ab else ''
                    table1 += f'<td{ab_css}>{ws}</td><td{ab_css}>{rc}</td><td{ab_css}>{cr}</td><td{ab_css}>{uf}</td>'
                    table1 += f'<td{ab_css} style="text-align:left; max-width:300px; white-space:normal;">{rem}</td>'

                table1 += "</tr>"

            if total_row1 is not None:
                table1 += "<tr style='background:#b8cce4;'>"
                for col in columns_order:
                    val = total_row1[col] if col in total_row1 else ""
                    align = ' style="text-align:left"' if col == "Name" else ''
                    table1 += f'<td{align}><b>{val}</b></td>'
                if uploaded_file_2 is not None:
                    table1 += f'<td><b>{"NA" if ws_total1 == 0 else ws_total1}</b></td>'
                    table1 += f'<td><b>{"NA" if rc_total1 == 0 else rc_total1}</b></td>'
                    table1 += f'<td><b>{"NA" if cr_total1 == 0 else cr_total1}</b></td>'
                    table1 += f'<td><b>{"NA" if uf_total1 == 0 else uf_total1}</b></td>'
                    table1 += f'<td><b>NA</b></td>'
                table1 += "</tr>"
            table1 += "</tbody></table>"

            # ===== TABLE 2 — NO COMPARISON =====
            table2 = build_header("table-daily2")
            total_row2 = None
            ws_total2 = cr_total2 = rc_total2 = uf_total2 = 0

            for _, row in day_df.iterrows():
                name = str(row["Name"]).strip()
                if name == "Total":
                    total_row2 = row
                    continue

                is_ab = str(row.get("Daily Cases", "")).strip() == "AB"
                row_style = " style='background:#e9dcc2;'" if is_ab else ""
                table2 += f"<tr{row_style}>"

                for col in columns_order:
                    val = row[col] if col in row else "NA"
                    if str(val) == "AB":
                        css = ' class="ab-cell2"'
                    else:
                        css = ''
                    align = ' style="text-align:left"' if col == "Name" else ''
                    table2 += f'<td{css}{align}>{val}</td>'

                if uploaded_file_2 is not None:
                    if is_ab:
                        ws = rc = cr = uf = rem = "AB"
                    else:
                        ws  = get_val(website_searching, name)
                        rc  = get_val(remark_checking, name)
                        cr  = get_val(credential_making, name)
                        uf  = get_val(upi_fraud, name)
                        rem = get_val(remark_col, name)

                        ws = "NA" if ws == 0 or ws == "NA" else ws
                        rc = "NA" if rc == 0 or rc == "NA" else rc
                        cr = "NA" if cr == 0 or cr == "NA" else cr
                        uf = "NA" if uf == 0 or uf == "NA" else uf

                    if str(ws) not in ("NA", "AB"): ws_total2 += int(ws)
                    if str(rc) not in ("NA", "AB"): rc_total2 += int(rc)
                    if str(cr) not in ("NA", "AB"): cr_total2 += int(cr)
                    if str(uf) not in ("NA", "AB"): uf_total2 += int(uf)

                    ab_css2 = ' class="ab-cell2"' if is_ab else ''
                    table2 += f'<td{ab_css2}>{ws}</td><td{ab_css2}>{rc}</td><td{ab_css2}>{cr}</td><td{ab_css2}>{uf}</td>'
                    table2 += f'<td{ab_css2} style="text-align:left; max-width:300px; white-space:normal;">{rem}</td>'

                table2 += "</tr>"

            if total_row2 is not None:
                table2 += "<tr style='background:#b8cce4;'>"
                for col in columns_order:
                    val = total_row2[col] if col in total_row2 else ""
                    align = ' style="text-align:left"' if col == "Name" else ''
                    table2 += f'<td{align}><b>{val}</b></td>'
                if uploaded_file_2 is not None:
                    table2 += f'<td><b>{"NA" if ws_total2 == 0 else ws_total2}</b></td>'
                    table2 += f'<td><b>{"NA" if rc_total2 == 0 else rc_total2}</b></td>'
                    table2 += f'<td><b>{"NA" if cr_total2 == 0 else cr_total2}</b></td>'
                    table2 += f'<td><b>{"NA" if uf_total2 == 0 else uf_total2}</b></td>'
                    table2 += f'<td><b>NA</b></td>'
                table2 += "</tr>"
            table2 += "</tbody></table>"

            # ===== RENDER =====
            st.markdown("#### 📊 Incorrect Summary Filled Data")
            wrapped1, h1 = with_copy_and_edit_button(shared_css + table1, table_id="daily-incorrect-table", height=550)
            components.html(wrapped1, height=h1, scrolling=True)

            st.markdown("#### 📋 Correct Summary")
            wrapped2, h2 = with_copy_and_edit_button(shared_css + table2, table_id="daily-correct-table", height=550)
            components.html(wrapped2, height=h2, scrolling=True)

            if mismatches and uploaded_file_2 is not None:
                st.warning(f"⚠️ {len(mismatches)} mismatches found — highlighted in red.")
            elif uploaded_file_2 is not None:
                st.success("✅ All values match with second file!")
        else:
            st.info("No daily summary data available.")

    # Excel export — now includes Daily Summary sheet
    excel_data = build_excel(summary_df, multiple_summary_df, freelancer_summary_df, daily_summary_df, crypto_summary_df, notFound_summary_df, website_searching=website_searching,
    credential_making=credential_making,
    remark_checking=remark_checking,        
    upi_fraud=upi_fraud,                   
    remark_col=remark_col )

    st.download_button(
        "📥 Download Summary Excel",
        data=excel_data,
        file_name="complete_summary.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("📤 Please upload a file to generate the report.")